#!/usr/bin/env python3
"""
Trend Micro WFBSS - Report client con sistema operativo Server in stato offline.

Flusso:
1. GET /SMPI/service/wfbss/api/customers
2. Per ogni customer: GET /SMPI/service/wfbss/api/components?cids=<CUSTOMER_ID>&page=N&limit=100
3. Filtra solo i computer con:
   - online == false  (stato offline)
   - platform contenente "server" (case-insensitive) → OS Windows Server
4. Genera un Excel per cliente (in memoria) con il dettaglio dei server offline
5. Invia il report ai referenti configurati in referenti.json

Esempio:
  python3 Server_Offline_WF.py
  python3 Server_Offline_WF.py --no-email       # genera Excel senza inviare mail
  python3 Server_Offline_WF.py --input-json raw.json   # test locale

Dipendenze:
  pip install requests python-dotenv openpyxl
"""

from __future__ import annotations

import argparse
import base64
import datetime as dt
import hashlib
import hmac
import json
import logging
import logging.handlers
import os
import re
import sys
import time
import uuid
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple
from urllib.parse import urlencode
from zoneinfo import ZoneInfo

import io
import smtplib
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import requests
from dotenv import load_dotenv

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
except ImportError as exc:  # pragma: no cover
    raise SystemExit(
        "Modulo openpyxl non installato. Installa le dipendenze con: "
        "pip install requests python-dotenv openpyxl"
    ) from exc


# ---------------------------------------------------------------------------
# Costanti
# ---------------------------------------------------------------------------

DEFAULT_ENV_PATH   = "/srv/Progetti_Pyhton/Server_Offline_WF/.Server_Offline_WF.env"
DEFAULT_OUTPUT_DIR = "/srv/Progetti_Pyhton/Server_Offline_WF"
DEFAULT_TIMEOUT    = 60
DEFAULT_LIMIT      = 100
DEFAULT_LOG_FILE   = "/var/log/Server_Offline_WF.log"

ADMIN_EMAIL = "magostini@euroinformatica.net"

SMPI_CUSTOMERS_PATH  = "/SMPI/service/wfbss/api/customers"
SMPI_COMPONENTS_PATH = "/SMPI/service/wfbss/api/components"

ROME_TZ = ZoneInfo("Europe/Rome")

# Parola chiave per identificare i sistemi operativi server nel campo platform.
# L'API WFBSS restituisce stringhe tipo "Windows Server 2019", "Windows Server 2022", ecc.
SERVER_PLATFORM_KEYWORD = "server"


# ---------------------------------------------------------------------------
# Errori di dominio
# ---------------------------------------------------------------------------

class SMPIError(RuntimeError):
    pass


# ---------------------------------------------------------------------------
# Modelli dati
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class SmtpConfig:
    host: str
    port: int
    user: str
    password: str
    from_addr: str
    from_name: str
    use_starttls: bool
    cc_internal: List[str]
    subject_prefix: str


@dataclass(frozen=True)
class AppConfig:
    env_file: str
    smpi_base_url: str
    smpi_token: Optional[str]
    smpi_secret: Optional[str]
    timeout: int
    output_dir: str
    log_file: str
    limit: int
    dump_raw: bool
    debug: bool
    verbose: bool
    stop_on_error: bool
    api_sleep: float
    email_json_file: str
    smtp: SmtpConfig


@dataclass(frozen=True)
class CustomerInfo:
    customer_id: str
    customer_name: str
    raw: Dict[str, Any] = field(default_factory=dict)


@dataclass
class CustomerComponents:
    customer_id: str
    customer_name: str
    computers: List[Dict[str, Any]] = field(default_factory=list)
    error: str = ""


@dataclass(frozen=True)
class ServerOfflineRow:
    customer_id: str
    customer_name: str
    computer_id: str
    computer_name: str
    ip_addr: str
    platform: str
    last_connect: str       # last_connect / last_seen come stringa ISO, se disponibile


# ---------------------------------------------------------------------------
# Logging e helper generici
# ---------------------------------------------------------------------------

def _setup_logging(log_file: str, verbose: bool = False) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    fmt = "%(asctime)s %(levelname)s %(message)s"

    root = logging.getLogger()
    root.setLevel(level)
    root.handlers.clear()

    sh = logging.StreamHandler(sys.stdout)
    sh.setLevel(level)
    sh.setFormatter(logging.Formatter(fmt))
    root.addHandler(sh)

    try:
        fh = logging.handlers.RotatingFileHandler(
            log_file, maxBytes=5 * 1024 * 1024, backupCount=5, encoding="utf-8"
        )
        fh.setLevel(level)
        fh.setFormatter(logging.Formatter(fmt))
        root.addHandler(fh)
    except (PermissionError, OSError):
        logging.warning("Impossibile scrivere su %s: log solo su stdout.", log_file)


def _none_if_empty(v: Optional[str]) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s or None


def _require(v: Optional[str], label: str) -> str:
    if v:
        return v
    raise SystemExit(f"Parametro obbligatorio mancante: {label}")


def _to_int(v: Any, *, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(v)
    except (TypeError, ValueError):
        try:
            return int(float(str(v).replace(",", ".")))
        except (TypeError, ValueError):
            return default


def _to_bool(v: Any) -> Optional[bool]:
    """Converte valori API eterogenei in bool. Ritorna None se il campo manca."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return v != 0
    s = str(v).strip().lower()
    if s in {"true", "1", "yes", "y", "online"}:
        return True
    if s in {"false", "0", "no", "n", "offline"}:
        return False
    return None


def _safe_filename(v: str) -> str:
    cleaned = "".join(c if c.isalnum() or c in {"-", "_", "."} else "_" for c in v)
    return cleaned.strip("_") or "file"


def _ensure_dir(p: str) -> Path:
    path = Path(p)
    path.mkdir(parents=True, exist_ok=True)
    return path


def _extract_list(payload: Any, keys: Sequence[str]) -> List[Any]:
    if isinstance(payload, list):
        return payload
    if not isinstance(payload, dict):
        return []
    for key in keys:
        value = payload.get(key)
        if isinstance(value, list):
            return value
    return []


# ---------------------------------------------------------------------------
# Client SMPI  (identico a Stato_Aggiornamenti_WF.py)
# ---------------------------------------------------------------------------

class SMPIClient:
    def __init__(self, base_url: str, access_token: str, secret_key: str,
                 timeout: int = DEFAULT_TIMEOUT) -> None:
        self.base_url = str(base_url).rstrip("/")
        self.access_token = access_token
        self.secret_key = secret_key
        self.timeout = timeout
        self.session = requests.Session()

    @staticmethod
    def _md5_b64(data: bytes) -> str:
        return base64.b64encode(hashlib.md5(data).digest()).decode()

    def _sign(self, posix: int, method: str, uri: str, body: bytes) -> str:
        msg = f"{posix}{method.upper()}{uri}"
        if body:
            msg += self._md5_b64(body)
        digest = hmac.new(self.secret_key.encode(), msg.encode(), hashlib.sha256).digest()
        return base64.b64encode(digest).decode()

    def request(self, method: str, path: str, *, payload: Optional[Dict[str, Any]] = None) -> Any:
        body = (
            json.dumps(payload, ensure_ascii=False, separators=(",", ":")).encode()
            if payload else b""
        )
        posix = int(dt.datetime.now(dt.timezone.utc).timestamp())
        method_u = method.upper()
        headers = {
            "x-access-token": self.access_token,
            "x-posix-time": str(posix),
            "x-signature": self._sign(posix, method_u, path, body),
            "x-traceid": str(uuid.uuid4()),
            "content-type": "application/json; charset=utf-8",
        }
        logging.debug("Chiamata API %s %s", method_u, path)
        resp = self.session.request(
            method=method_u,
            url=f"{self.base_url}{path}",
            data=body or None,
            headers=headers,
            timeout=self.timeout,
        )
        if resp.status_code >= 400:
            raise SMPIError(f"HTTP {resp.status_code} su {method_u} {path}: {resp.text[:800]}")
        if not resp.text.strip():
            return {}
        try:
            return resp.json()
        except ValueError as exc:
            raise SMPIError(f"JSON non valido da {path}: {resp.text[:400]}") from exc

    def request_safe(self, method: str, path: str, *, payload: Optional[Dict[str, Any]] = None) -> Tuple[bool, Any]:
        try:
            return True, self.request(method, path, payload=payload)
        except SMPIError as exc:
            return False, str(exc)

    def list_customers(self) -> List[Dict[str, Any]]:
        """GET /customers con paginazione offset/limit e fallback senza parametri."""
        all_records: List[Any] = []
        offset, limit = 0, 200
        while True:
            path = f"{SMPI_CUSTOMERS_PATH}?offset={offset}&limit={limit}"
            ok, payload = self.request_safe("GET", path)
            if not ok:
                if offset == 0:
                    payload = self.request("GET", SMPI_CUSTOMERS_PATH)
                    records = _extract_list(payload, ("customers", "data", "items", "result"))
                    return [r for r in records if isinstance(r, dict)]
                break
            page = _extract_list(payload, ("customers", "data", "items", "result"))
            if not page:
                break
            all_records.extend(page)
            if len(page) < limit:
                break
            offset += limit
        if all_records:
            return [r for r in all_records if isinstance(r, dict)]

        payload = self.request("GET", SMPI_CUSTOMERS_PATH)
        records = _extract_list(payload, ("customers", "data", "items", "result"))
        return [r for r in records if isinstance(r, dict)]

    def list_components_for_customer(self, customer_id: str, *, limit: int = DEFAULT_LIMIT) -> List[Dict[str, Any]]:
        """GET /components per un singolo customer, con paging page/limit."""
        page = 1
        computers: List[Dict[str, Any]] = []
        limit = max(1, min(int(limit), 100))

        while True:
            query = urlencode({"cids": customer_id, "page": page, "limit": limit})
            path = f"{SMPI_COMPONENTS_PATH}?{query}"
            payload = self.request("GET", path)
            customers_page = _extract_list(payload, ("customers", "data", "items", "result"))

            page_computers: List[Dict[str, Any]] = []
            for customer in customers_page:
                if not isinstance(customer, dict):
                    continue
                for comp in customer.get("computers") or []:
                    if isinstance(comp, dict):
                        page_computers.append(comp)

            computers.extend(page_computers)
            paging = payload.get("paging") if isinstance(payload, dict) else {}
            total = _to_int((paging or {}).get("total"), default=0)

            logging.debug(
                "components customer=%s page=%s records=%s total=%s",
                customer_id, page, len(page_computers), total or "n/a",
            )

            if total and page * limit >= total:
                break
            if not page_computers or len(page_computers) < limit:
                break
            page += 1

        return computers


# ---------------------------------------------------------------------------
# Parsing customers/components
# ---------------------------------------------------------------------------

def _customer_id(rec: Dict[str, Any]) -> str:
    return str(rec.get("id") or rec.get("customer_id") or rec.get("cid") or "").strip()


def _customer_name(rec: Dict[str, Any], fallback_id: str = "") -> str:
    name = (
        rec.get("name") or rec.get("customer") or rec.get("customer_name")
        or rec.get("company") or rec.get("company_name") or fallback_id or "<sconosciuto>"
    )
    return str(name).strip() or fallback_id or "<sconosciuto>"


def build_customer_infos(records: Iterable[Dict[str, Any]]) -> List[CustomerInfo]:
    infos: List[CustomerInfo] = []
    seen: set[str] = set()
    for rec in records:
        if not isinstance(rec, dict):
            continue
        cid = _customer_id(rec)
        if not cid or cid.upper() in seen:
            continue
        seen.add(cid.upper())
        infos.append(CustomerInfo(customer_id=cid, customer_name=_customer_name(rec, cid), raw=rec))
    infos.sort(key=lambda c: c.customer_name.lower())
    return infos


def fetch_components(config: AppConfig) -> List[CustomerComponents]:
    token  = _require(config.smpi_token,   "SMPI_ACCESS_TOKEN")
    secret = _require(config.smpi_secret,  "SMPI_SECRET_KEY")
    base_url = _require(config.smpi_base_url, "SMPI_BASE_URL")

    client = SMPIClient(base_url, token, secret, config.timeout)

    logging.info("Recupero clienti da %s ...", SMPI_CUSTOMERS_PATH)
    customer_records = client.list_customers()
    customers = build_customer_infos(customer_records)
    logging.info("Clienti trovati: %d", len(customers))

    result: List[CustomerComponents] = []
    raw_dump: Dict[str, Any] = {"customers": customer_records, "components": []}

    for idx, cust in enumerate(customers, start=1):
        logging.info("[%d/%d] Recupero componenti: %s", idx, len(customers), cust.customer_name)
        try:
            computers = client.list_components_for_customer(cust.customer_id, limit=config.limit)
            result.append(CustomerComponents(cust.customer_id, cust.customer_name, computers))
            raw_dump["components"].append({
                "customer_id": cust.customer_id,
                "customer_name": cust.customer_name,
                "computers": computers,
            })
            logging.info("  PC ricevuti: %d", len(computers))
        except SMPIError as exc:
            detail = str(exc)
            logging.error("Errore componenti per %s: %s", cust.customer_name, detail)
            result.append(CustomerComponents(cust.customer_id, cust.customer_name, [], detail))
            if config.stop_on_error:
                raise
        if config.api_sleep > 0 and idx < len(customers):
            time.sleep(config.api_sleep)

    if config.dump_raw:
        out_dir = _ensure_dir(config.output_dir)
        ts = dt.datetime.now(ROME_TZ).strftime("%Y%m%d_%H%M%S")
        raw_path = out_dir / f"raw_wfbss_components_{ts}.json"
        raw_path.write_text(json.dumps(raw_dump, ensure_ascii=False, indent=2), encoding="utf-8")
        logging.info("Dump raw salvato: %s", raw_path)

    return result


def load_components_from_json(path: str) -> List[CustomerComponents]:
    """
    Carica un JSON locale per test/debug.

    Formati accettati:
    - {"customers": [{"id": "...", "name": "...", "computers": [...]}, ...]}
    - [{"id": "...", "name": "...", "computers": [...]}, ...]
    - dump prodotto da --dump-raw: {"components": [{"customer_id": "...", ...}]}
    """
    payload = json.loads(Path(path).read_text(encoding="utf-8"))

    if isinstance(payload, dict) and isinstance(payload.get("components"), list):
        records = payload["components"]
    elif isinstance(payload, dict) and isinstance(payload.get("customers"), list):
        records = payload["customers"]
    elif isinstance(payload, list):
        records = payload
    else:
        raise SystemExit(f"Formato JSON non riconosciuto: {path}")

    result: List[CustomerComponents] = []
    for rec in records:
        if not isinstance(rec, dict):
            continue
        cid  = str(rec.get("customer_id") or rec.get("id") or rec.get("cid") or "").strip()
        name = str(rec.get("customer_name") or rec.get("name") or rec.get("customer") or cid or "<sconosciuto>").strip()
        computers = rec.get("computers") or []
        result.append(CustomerComponents(cid, name, [c for c in computers if isinstance(c, dict)]))
    return result


# ---------------------------------------------------------------------------
# Filtri: offline + OS server
# ---------------------------------------------------------------------------

def _is_computer_offline(computer: Dict[str, Any]) -> bool:
    """
    True se il computer risulta offline.

    Logica speculare a _is_computer_online() dello script pattern:
    - campo 'online' booleano → usa direttamente
    - fallback: status != 1  (1 = online per Windows/Mac in WFBSS)
    """
    online = _to_bool(computer.get("online"))
    if online is not None:
        return not online
    status = _to_int(computer.get("status"), default=0)
    return status != 1


def _is_server_os(computer: Dict[str, Any]) -> bool:
    """
    True se il campo platform/os contiene la parola 'server' (case-insensitive).
    Esempi di stringhe restituite dall'API: 'Windows Server 2019', 'Windows Server 2022',
    'Windows Server 2016', 'Windows Server 2012 R2'.
    """
    platform = str(computer.get("platform") or computer.get("os") or "").lower()
    return SERVER_PLATFORM_KEYWORD in platform


def _last_connect(computer: Dict[str, Any]) -> str:
    """
    Restituisce il timestamp dell'ultimo collegamento come stringa.
    Campi tentati (in ordine di priorità): last_connect, last_seen, last_update.
    """
    for key in ("last_connect", "last_seen", "last_update", "lastConnect", "lastSeen"):
        val = computer.get(key)
        if val:
            return str(val).strip()
    return ""


# ---------------------------------------------------------------------------
# Raccolta righe server offline
# ---------------------------------------------------------------------------

def collect_server_offline_rows(customer_components: Iterable[CustomerComponents]) -> List[ServerOfflineRow]:
    rows: List[ServerOfflineRow] = []
    skipped_online = skipped_not_server = 0

    for cust in customer_components:
        for computer in cust.computers:
            if not _is_computer_offline(computer):
                skipped_online += 1
                continue
            if not _is_server_os(computer):
                skipped_not_server += 1
                continue
            rows.append(ServerOfflineRow(
                customer_id=cust.customer_id,
                customer_name=cust.customer_name,
                computer_id=str(computer.get("id") or computer.get("ccid") or "").strip(),
                computer_name=str(computer.get("name") or "<senza_nome>").strip(),
                ip_addr=str(computer.get("ip_addr") or computer.get("ip") or "").strip(),
                platform=str(computer.get("platform") or computer.get("os") or "").strip(),
                last_connect=_last_connect(computer),
            ))

    logging.info(
        "Server offline rilevati: %d (online saltati=%d, non-server saltati=%d)",
        len(rows), skipped_online, skipped_not_server,
    )
    return rows


# ---------------------------------------------------------------------------
# Stile Excel  (palette Euro Informatica identica allo script pattern)
# ---------------------------------------------------------------------------

INVALID_SHEET_CHARS = set('[]:*?/\\')


def _safe_sheet_name(name: str) -> str:
    cleaned = "".join("_" if c in INVALID_SHEET_CHARS else c for c in str(name or "Sheet"))
    cleaned = cleaned.strip().strip("'") or "Sheet"
    return cleaned[:31]


def _set_widths(ws, widths: Dict[str, float]) -> None:
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def _add_table(ws, table_name: str, last_col: str, last_row: int) -> None:
    if last_row < 2:
        ws.auto_filter.ref = f"A1:{last_col}1"
        return
    ref = f"A1:{last_col}{last_row}"
    table = Table(displayName=table_name, ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    ws.add_table(table)


# ---------------------------------------------------------------------------
# Generazione workbook cliente  (in memoria, senza scrittura su disco)
# ---------------------------------------------------------------------------

def write_client_workbook(
    *,
    rows: Sequence[ServerOfflineRow],
    customer_name: str,
    timestamp: dt.datetime,
) -> Tuple[bytes, str]:
    """
    Genera un file Excel per un singolo cliente con i server offline.
    Restituisce (bytes_excel, nome_file) senza scrivere nulla su disco.

    Struttura:
      - Foglio 1: "Server offline"   – lista dei server con OS, IP, ultimo collegamento
    """
    safe     = _safe_filename(customer_name)[:60]
    filename = f"wfbss_server_offline_{safe}_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx"

    ROW_FONT    = Font(name="Arial", size=10)
    HDR_FILL    = PatternFill("solid", fgColor="1F0034")
    HDR_FONT    = Font(name="Arial", color="FFFFFF", bold=True, size=10)
    THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))
    OFFLINE_FILL = PatternFill("solid", fgColor="FFD966")   # giallo — server irraggiungibile

    def _hdr(ws) -> None:
        for cell in ws[1]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 20

    def _dc(cell, fill=None) -> None:
        cell.font = ROW_FONT
        cell.alignment = Alignment(vertical="center")
        cell.border = THIN_BORDER
        if fill:
            cell.fill = fill

    wb = Workbook()

    # ------------------------------------------------------------------
    # Foglio 1 – Server offline
    # ------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Server offline"
    ws1.freeze_panes = "A2"
    ws1.append(["Nome server", "IP", "Sistema operativo", "Ultimo collegamento"])
    _hdr(ws1)

    for r in sorted(rows, key=lambda x: x.computer_name.lower()):
        ws1.append([r.computer_name, r.ip_addr, r.platform, r.last_connect])
        for col in range(1, 5):
            _dc(ws1.cell(row=ws1.max_row, column=col), OFFLINE_FILL)

    _set_widths(ws1, {"A": 30, "B": 18, "C": 36, "D": 26})
    if rows:
        _add_table(ws1, "Server_offline", "D", ws1.max_row)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue(), filename


# ---------------------------------------------------------------------------
# Referenti e invio mail
# ---------------------------------------------------------------------------

def load_referenti(path: str) -> Dict[str, List[str]]:
    """
    Carica il file referenti.json.
    Formato atteso:
        {"NOME CLIENTE": ["email1@esempio.it", "email2@esempio.it"], ...}
    """
    try:
        raw: Dict[str, Any] = json.loads(Path(path).read_text(encoding="utf-8"))
    except FileNotFoundError:
        logging.warning("File referenti non trovato: %s — nessuna mail verrà inviata.", path)
        return {}
    except json.JSONDecodeError as exc:
        logging.error("referenti.json non valido: %s", exc)
        return {}

    result: Dict[str, List[str]] = {}
    for name, emails in raw.items():
        key = str(name).strip()
        if not key:
            continue
        valid = [e for e in (emails or []) if isinstance(e, str) and "@" in e]
        if valid:
            result[key] = valid
        else:
            logging.warning("Referenti: nessuna email valida per cliente '%s'.", key)
    return result


def _referenti_lookup(referenti: Dict[str, List[str]], customer_name: str) -> List[str]:
    """Lookup case-insensitive + strip sul nome cliente."""
    key = customer_name.strip()
    if key in referenti:
        return referenti[key]
    key_lower = key.lower()
    for k, v in referenti.items():
        if k.lower() == key_lower:
            return v
    return []


def _smtp_from_header(cfg: SmtpConfig) -> str:
    if cfg.from_name:
        return f"{cfg.from_name} <{cfg.from_addr}>"
    return cfg.from_addr


def _all_recipients(msg: MIMEMultipart) -> List[str]:
    recipients: List[str] = []
    for header in ("To", "Bcc"):
        val = msg.get(header, "")
        if val:
            recipients.extend(e.strip() for e in val.split(",") if e.strip())
    return recipients


def _send_smtp(smtp_cfg: SmtpConfig, msg: MIMEMultipart) -> None:
    """Invia via SMTP (plain o STARTTLS). Supporta relay senza AUTH."""
    conn = smtplib.SMTP(smtp_cfg.host, smtp_cfg.port, timeout=30)
    try:
        conn.ehlo()
        if smtp_cfg.use_starttls:
            conn.starttls()
            conn.ehlo()
        if smtp_cfg.user and smtp_cfg.password:
            server_supports_auth = bool(conn.esmtp_features.get("auth"))
            if server_supports_auth:
                conn.login(smtp_cfg.user, smtp_cfg.password)
            else:
                logging.debug(
                    "SMTP: server %s:%s non dichiara AUTH — login saltato (relay interno).",
                    smtp_cfg.host, smtp_cfg.port,
                )
        conn.sendmail(smtp_cfg.from_addr, _all_recipients(msg), msg.as_string())
    finally:
        try:
            conn.quit()
        except Exception:
            pass


def send_client_report(
    smtp_cfg: SmtpConfig,
    customer_name: str,
    to_emails: List[str],
    excel_data: bytes,
    excel_filename: str,
    timestamp: dt.datetime,
    n_server: int,
) -> None:
    """Invia il report Excel (in memoria) di un cliente ai suoi referenti, con Bcc interni."""
    msg = MIMEMultipart()
    msg["From"]    = _smtp_from_header(smtp_cfg)
    msg["To"]      = ", ".join(to_emails)
    if smtp_cfg.cc_internal:
        msg["Bcc"] = ", ".join(smtp_cfg.cc_internal)
    msg["Subject"] = (
        f"{smtp_cfg.subject_prefix}{customer_name} — {timestamp.strftime('%d/%m/%Y')}"
    )
    body = f"""
    <html>
    <body style="font-family: Arial, sans-serif; font-size: 14px; color: #000000;">
        <p>Gentile cliente,</p>

        <p>
        Dal monitoraggio del tenant Trend Micro Worry-Free Business Security rileviamo
        <strong>{n_server} server</strong> con sistema operativo Windows Server
        attualmente in stato <strong>offline</strong> sul tenant {customer_name}.
        </p>

        <p>
        In allegato trovate il report Excel con il dettaglio dei dispositivi non raggiungibili,
        comprensivo di IP, sistema operativo e data dell'ultimo collegamento registrato.
        </p>

        <p>
        Vi invitiamo a verificare lo stato di questi sistemi il prima possibile al fine di
        garantire la continuità della protezione antivirus e la corretta gestione delle policy
        di sicurezza.
        </p>

        <p>
        Report generato il {timestamp.strftime('%d/%m/%Y alle %H:%M')}.
        </p>

        <p>
        Restiamo a disposizione per qualsiasi chiarimento o per supportarvi nelle attività.
        </p>

        <p>
        Cordiali Saluti<br><br>
        Euro Informatica S.p.A.
        </p>

        <hr style="border:none;border-top:1px solid #d9d9d9;margin-top:24px;"/>
        <p style="font-size:11px;color:#777;">Notifica generata automaticamente dal sistema
        di monitoraggio.</p>
    </body>
    </html>
    """
    msg.attach(MIMEText(body, "html", "utf-8"))

    part = MIMEApplication(excel_data, Name=excel_filename)
    part["Content-Disposition"] = f'attachment; filename="{excel_filename}"'
    msg.attach(part)

    _send_smtp(smtp_cfg, msg)
    logging.info(
        "Mail inviata: %s → To:%s Bcc:%s (%d server segnalati)",
        customer_name, ", ".join(to_emails),
        ", ".join(smtp_cfg.cc_internal) or "-", n_server,
    )


def send_missing_referenti_alert(
    smtp_cfg: SmtpConfig,
    missing: List[str],
    timestamp: dt.datetime,
) -> None:
    """
    Invia a ADMIN_EMAIL l'elenco dei clienti con server offline
    per i quali non è stato trovato alcun referente in referenti.json.
    """
    if not missing:
        return
    msg = MIMEMultipart()
    msg["From"]    = _smtp_from_header(smtp_cfg)
    msg["To"]      = ADMIN_EMAIL
    msg["Subject"] = (
        f"{smtp_cfg.subject_prefix}referenti mancanti — {timestamp.strftime('%d/%m/%Y')}"
    )
    elenco = "\n".join(f"  - {name}" for name in sorted(missing))
    body = (
        f"I seguenti clienti hanno server Windows con stato offline\n"
        f"ma non hanno referenti configurati in referenti.json:\n\n"
        f"{elenco}\n\n"
        f"Aggiornare il file referenti.json per abilitare l'invio automatico del report.\n\n"
        f"Rilevazione del {timestamp.strftime('%d/%m/%Y alle %H:%M')}.\n\n"
        f"Euro Informatica S.p.A.\n"
    )
    msg.attach(MIMEText(body, "plain", "utf-8"))
    _send_smtp(smtp_cfg, msg)
    logging.info(
        "Alert referenti mancanti inviato a %s per %d clienti: %s",
        ADMIN_EMAIL, len(missing), ", ".join(sorted(missing)),
    )


# ---------------------------------------------------------------------------
# CLI e orchestrazione
# ---------------------------------------------------------------------------

def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    pre = argparse.ArgumentParser(add_help=False)
    pre.add_argument("--env-file", default=os.getenv("WF_SERVER_OFFLINE_ENV_FILE", DEFAULT_ENV_PATH))
    pre_args, _ = pre.parse_known_args(argv)
    if pre_args.env_file and Path(pre_args.env_file).exists():
        load_dotenv(pre_args.env_file, override=False)

    p = argparse.ArgumentParser(
        description="Trend Micro WFBSS - Report server (OS Server) in stato offline."
    )
    p.add_argument("--env-file",        default=pre_args.env_file)
    p.add_argument("--smpi-base-url",   default=os.getenv("SMPI_BASE_URL", os.getenv("LMPI_BASE_URL", "")))
    p.add_argument("--smpi-token",      default=os.getenv("SMPI_ACCESS_TOKEN", os.getenv("LMPI_ACCESS_TOKEN")))
    p.add_argument("--smpi-secret",     default=os.getenv("SMPI_SECRET_KEY", os.getenv("LMPI_SECRET_KEY")))
    p.add_argument("--timeout",         type=int, default=int(os.getenv("SMPI_TIMEOUT", str(DEFAULT_TIMEOUT))))
    p.add_argument("--limit",           type=int, default=int(os.getenv("WF_SERVER_OFFLINE_LIMIT", str(DEFAULT_LIMIT))),
                   help="Record per pagina per /components, max 100.")
    p.add_argument("--output-dir",      default=os.getenv("WF_SERVER_OFFLINE_OUTPUT_DIR", os.getenv("LMPI_OUTPUT_DIR", DEFAULT_OUTPUT_DIR)))
    p.add_argument("--log-file",        default=os.getenv("WF_SERVER_OFFLINE_LOG_FILE", DEFAULT_LOG_FILE))
    p.add_argument("--input-json",      help="Usa un JSON locale invece delle API, utile per test/debug.")
    p.add_argument("--dump-raw",        action="store_true", help="Salva JSON grezzo in output-dir.")
    p.add_argument("--debug",           action="store_true", help="Stampa dettagli aggiuntivi.")
    p.add_argument("--verbose",         action="store_true", help="Logging DEBUG.")
    p.add_argument("--stop-on-error",   action="store_true", help="Interrompe al primo errore API su un cliente.")
    p.add_argument("--api-sleep",       type=float, default=float(os.getenv("WF_SERVER_OFFLINE_API_SLEEP", "0")),
                   help="Pausa in secondi tra una chiamata cliente e la successiva.")
    # SMTP — variabili allineate al .env del progetto
    p.add_argument("--smtp-host",       default=os.getenv("SMTP_HOST", ""))
    p.add_argument("--smtp-port",       type=int, default=int(os.getenv("SMTP_PORT", "25")))
    p.add_argument("--smtp-user",       default=os.getenv("SMTP_USER", ""))
    p.add_argument("--smtp-password",   default=os.getenv("SMTP_PASSWORD", ""))
    p.add_argument("--smtp-from",       default=os.getenv("MAIL_FROM", os.getenv("SMTP_USER", "")))
    p.add_argument("--smtp-from-name",  default=os.getenv("MAIL_FROM_NAME", "Notification - Euro Informatica"))
    p.add_argument("--smtp-starttls",   action="store_true",
                   default=os.getenv("SMTP_STARTTLS", "false").lower() == "true")
    p.add_argument("--mail-cc-internal",    default=os.getenv("MAIL_CC_INTERNAL", ""))
    p.add_argument("--mail-subject-prefix", default=os.getenv("MAIL_SUBJECT_PREFIX", "[Trend Micro WFBSS] Server offline - "))
    p.add_argument("--email-json-file", default=os.getenv("EMAIL_TO_JSON_FILE", ""))
    p.add_argument("--no-email",        action="store_true",
                   help="Genera i file Excel ma non invia alcuna mail (utile per test).")
    return p.parse_args(argv)


def build_config(args: argparse.Namespace) -> AppConfig:
    return AppConfig(
        env_file=args.env_file,
        smpi_base_url=str(args.smpi_base_url or "").rstrip("/"),
        smpi_token=_none_if_empty(args.smpi_token),
        smpi_secret=_none_if_empty(args.smpi_secret),
        timeout=args.timeout,
        output_dir=str(args.output_dir or DEFAULT_OUTPUT_DIR),
        log_file=str(args.log_file or DEFAULT_LOG_FILE),
        limit=max(1, min(int(args.limit), 100)),
        dump_raw=bool(args.dump_raw),
        debug=bool(args.debug),
        verbose=bool(args.verbose),
        stop_on_error=bool(args.stop_on_error),
        api_sleep=max(0.0, float(args.api_sleep or 0)),
        email_json_file=str(args.email_json_file or ""),
        smtp=SmtpConfig(
            host=str(args.smtp_host or ""),
            port=int(args.smtp_port or 25),
            user=str(args.smtp_user or ""),
            password=str(args.smtp_password or ""),
            from_addr=str(args.smtp_from or args.smtp_user or ""),
            from_name=str(args.smtp_from_name or ""),
            use_starttls=bool(args.smtp_starttls),
            cc_internal=[
                e.strip() for e in str(args.mail_cc_internal or "").split(",")
                if e.strip() and "@" in e
            ],
            subject_prefix=str(args.mail_subject_prefix or ""),
        ),
    )


def main(argv: Optional[List[str]] = None) -> int:
    args   = parse_args(argv)
    config = build_config(args)

    _ensure_dir(str(Path(config.log_file).parent))
    _setup_logging(config.log_file, verbose=config.verbose)

    if not config.smpi_base_url and not args.input_json:
        logging.error(
            "SMPI_BASE_URL non configurato e nessun --input-json specificato. "
            "Verifica il file .env o passa --smpi-base-url."
        )
        return 1

    ts = dt.datetime.now(ROME_TZ)
    logging.info(
        "Avvio @ %s | smpi=%s",
        ts.isoformat(timespec="seconds"),
        config.smpi_base_url or "<non configurato>",
    )

    # ------------------------------------------------------------------
    # Raccolta dati
    # ------------------------------------------------------------------
    if args.input_json:
        logging.info("Caricamento da JSON locale: %s", args.input_json)
        customer_components = load_components_from_json(args.input_json)
    else:
        customer_components = fetch_components(config)

    server_offline_rows = collect_server_offline_rows(customer_components)

    logging.info("Server offline con OS Server rilevati: %d", len(server_offline_rows))

    if config.debug:
        for row in server_offline_rows[:100]:
            logging.debug(
                "OFFLINE | %s | %s | %s | %s | last=%s",
                row.customer_name, row.computer_name, row.ip_addr,
                row.platform, row.last_connect or "n/d",
            )

    # ------------------------------------------------------------------
    # Raggruppa per cliente
    # ------------------------------------------------------------------
    rows_by_customer: Dict[str, List[ServerOfflineRow]] = defaultdict(list)
    for row in server_offline_rows:
        rows_by_customer[row.customer_id].append(row)

    # ------------------------------------------------------------------
    # Genera Excel per cliente e invia via mail
    # ------------------------------------------------------------------
    no_email: bool = getattr(args, "no_email", False)

    referenti: Dict[str, List[str]] = {}
    if config.email_json_file:
        referenti = load_referenti(config.email_json_file)
    else:
        logging.warning("EMAIL_TO_JSON_FILE non configurato: nessuna mail verrà inviata.")
        no_email = True

    n_clienti_con_offline: int = 0
    sent_ok:  List[str] = []
    sent_err: List[str] = []
    missing_referenti: List[str] = []

    for cust in sorted(customer_components, key=lambda c: c.customer_name.lower()):
        rows_out = rows_by_customer.get(cust.customer_id, [])
        if not rows_out:
            continue  # nessun server offline per questo cliente

        n_clienti_con_offline += 1

        excel_data, excel_filename = write_client_workbook(
            rows=rows_out,
            customer_name=cust.customer_name,
            timestamp=ts,
        )
        logging.info("Excel generato in memoria: %s (%d server)", excel_filename, len(rows_out))

        if no_email:
            continue

        emails = _referenti_lookup(referenti, cust.customer_name)
        if not emails:
            logging.warning(
                "Referente mancante per '%s': report non inviato.",
                cust.customer_name,
            )
            missing_referenti.append(cust.customer_name)
            continue

        try:
            send_client_report(
                smtp_cfg=config.smtp,
                customer_name=cust.customer_name,
                to_emails=emails,
                excel_data=excel_data,
                excel_filename=excel_filename,
                timestamp=ts,
                n_server=len(rows_out),
            )
            sent_ok.append(cust.customer_name)
        except Exception as exc:
            logging.error(
                "Errore invio mail per '%s' a %s: %s",
                cust.customer_name, emails, exc,
            )
            sent_err.append(cust.customer_name)

    # Invia alert per i clienti senza referente
    if missing_referenti and not no_email:
        try:
            send_missing_referenti_alert(
                smtp_cfg=config.smtp,
                missing=missing_referenti,
                timestamp=ts,
            )
        except Exception as exc:
            logging.error("Errore invio alert referenti mancanti: %s", exc)

    # ------------------------------------------------------------------
    # Riepilogo finale
    # ------------------------------------------------------------------
    print()
    print("=== Trend Micro WFBSS - Report server offline ===")
    print(f"Clienti analizzati            : {len(customer_components)}")
    print(f"Server offline (OS Server)    : {len(server_offline_rows)}")
    print(f"Clienti con server offline    : {n_clienti_con_offline}")
    if not no_email:
        print(f"Mail inviate OK               : {len(sent_ok)}")
        if sent_err:
            print(f"Mail con errore               : {len(sent_err)} — {', '.join(sent_err)}")
        if missing_referenti:
            print(f"Referenti mancanti            : {len(missing_referenti)} — alert inviato a {ADMIN_EMAIL}")
    print(f"Log                           : {config.log_file}")
    print()

    logging.info("Completato.")
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except KeyboardInterrupt:
        raise SystemExit("Interrotto dall'utente")
    except SMPIError as exc:
        logging.error("Errore SMPI: %s", exc)
        raise SystemExit(str(exc))